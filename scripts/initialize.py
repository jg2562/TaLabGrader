import pip
import sys
import csv
import grader.utils as utils
import openpyxl
from grader.browser.gradeBrowser import GradeBrowser

config = utils.load_json("./config/general.json")
workbook_name = config['spreadsheet']
try:
    assert sys.version_info >= (3,0)
except AssertionError:
    raise AssertionError("Must use python 3.0 or greater")

# Get user data for comments sheet
browser = GradeBrowser(config)
comments_data = browser.get_user_data()

# open and create grading sheet, check if rubric sheet exists
workbook = openpyxl.load_workbook(workbook_name)
workbook.active.title = "Rubric"
comment_sheet = workbook.create_sheet("Comments", 0)
comment_sheet.title = "Comments"
for row, name in enumerate(sorted(comments_data, key=lambda x: x[0])):
    comment_sheet.cell(row=row + 2, column=1).value = name[0]
    comment_sheet.cell(row=row + 2, column=2).value = name[1]
    comment_sheet.cell(row=row + 2, column=3).value = name[2]
    comment_sheet.cell(row=row + 2, column=4).value = name[3]

col = 1
for col_name in ["Last Name", "First Name","Username", "Student ID"]:
    comment_sheet.cell(row=1, column=col).value = col_name
    col += 1
for lab_num in range(12):
    comment_sheet.cell(row=1, column=col).value = "Lab " + str(lab_num + 1)
    col += 1

workbook.save(workbook_name)
workbook.close()
browser.close()
