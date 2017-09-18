import pip
import sys
import csv
import utils


username_file = 'usernames.json'
workbook_name = 'grade-sheet.xlsx'
try:
    assert sys.version_info >= (3,0)
except AssertionError:
    raise AssertionError("Must use python 3.0 or greater")


# download packages
pip.main(['install', 'openpyxl', '--user'])
pip.main(['install', 'splinter', '--user'])
pip.main(['install', 'request', '--user'])
pip.main(['install', 'pycodestyle', '--user'])


# set up usernames json file
usernames = {}
comments_data = []
with open('gradebook.csv') as csvfile:
    csvreader = csv.reader(csvfile)
    csvreader.__next__()
    for row in csvreader:
        usernames[row[1].strip() + " " + row[0].strip()] =  row[2]
        comments_data.append((row[1].strip(), row[0].strip(), row[2].strip(), row[3].strip()))
utils.save_json(username_file, usernames)


# open and create grading sheet, check if rubric sheet exists
import openpyxl
workbook = openpyxl.load_workbook(workbook_name)
workbook.active.title = "Rubric"
comment_sheet = workbook.create_sheet("Comments", 0)
comment_sheet.title = "Comments"
for row, name in enumerate(sorted(comments_data, key=lambda x: x[1])):
    comment_sheet.cell(row=row + 2, column=1).value = name[1]
    comment_sheet.cell(row=row + 2, column=2).value = name[0]
    comment_sheet.cell(row=row + 2, column=3).value = name[2]
    comment_sheet.cell(row=row + 2, column=4).value = name[3]

col = 1
for col_name in ["Last Name", "First Name","Username", "Student ID"]:
    comment_sheet.cell(row=1, column=col).value = col_name
    col += 1
for lab_num in range(12):
    comment_sheet.cell(row=1, column=col).value = "Lab " + str(lab_num + 1)
    col += 1

workbook.save(workbook_name)
workbook.close()
