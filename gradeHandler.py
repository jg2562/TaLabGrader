import openpyxl
import utils


class GradeHandler:
    def __init__(self, wb_name, grader_name, groups, usernames, lab_number):
        self._group_grades = {}
        self._grades = {}

        sheet = self._get_work_sheet(wb_name, lab_number)
        self._load_group_grades(sheet)
        self._load_student_grades(groups, usernames)
        self.grader = grader_name

    def _get_work_sheet(self, workbook_name, lab_number):
        wb = openpyxl.load_workbook(workbook_name, data_only=True)
        sheet = wb.get_sheet_by_name("Lab {}".format(lab_number))
        return sheet

    def _load_group_grades(self, sheet):
        row = 2
        group_name = sheet.cell(row=row, column=1).value
        while group_name:
            self._load_group_grade(sheet, row)
            row += 1
            group_name = sheet.cell(row=row, column=1).value

    def _load_group_grade(self, sheet, group_row_num):
        group_num = group_row_num - 2
        group_row = sheet[group_row_num][2:]
        group_row = [cell for cell in group_row if cell.value]
        self._group_grades[group_num] = {}
        self._group_grades[group_num]["Grade"] = str(sheet.cell(row=group_row_num, column=2).value)
        comments = []
        for cell in group_row:
            com = cell.comment
            sub_comment = ""
            if com:
                sub_comment = com.text
            point_diff = int(cell.value) - int(sheet[cell.column][0].value)
            if point_diff != 0:
                if not sub_comment:
                    raise ValueError("No note found on row {} and column {}".format(cell.row, cell.column))
                sub_comment = "{:+} : {}".format(point_diff, sub_comment)
                comments.append(sub_comment)
        self._group_grades[group_num]["Note"] = "\n".join(comments)

    def _load_student_grades(self, groups, usernames):
        rev_usernames = self._reverse_username_dict(usernames)
        for group in self._group_grades:
            for student in groups[str(group)]:
                student_key = rev_usernames[student]
                self._grades[student_key] = {}
                for category in self._group_grades[group]:
                    self._grades[student_key][category] = self._group_grades[group][category]

    def _reverse_username_dict(self, username):
        reverse = {}
        for key,value in username.items():
            reverse[value] = key
        return reverse

    def get_grade(self, name):
        return self._grades[name]["Grade"]

    def get_note(self, name):
        return self._grades[name]["Note"]

    def get_all_students(self):
        return self._grades.keys()


if __name__ == "__main__":
    usernames = utils.load_json("usernames.json")
    groups = utils.load_json("groups.json")
    handler = GradeHandler("./grading-sheet.xlsx", "Jack", groups, usernames, 10)
    grade = handler._grades
    for student in grade:
        print(student)
        print(grade[student]["Note"])
        print("----")
