import openpyxl
import grader.utils as utils

class GradeSheet():
    def __init__(self, config):
        self.workbook_name = config["spreadsheet"]
        self._active_sheet = None
        try:
            self.wb = openpyxl.load_workbook(self.workbook_name)
        except FileNotFoundError:
            self.wb = openpyxl.Workbook()

    def create_grade_sheet(self, groups, lab_config, group_grades):
        lab_config = lab_config["assignment"]
        lab_name = lab_config["lab name"]
        lab_sheet = self._get_lab_sheet(lab_name)
        self._section_ranges = self._get_section_ranges(lab_config["sections"])
        self._active_sheet = lab_sheet
        self._setup_header(lab_config)
        self._create_group_rows(lab_config, groups, group_grades)
        self.wb.save(self.workbook_name)

    def _get_lab_sheet(self, lab_name):
        if lab_name not in self.wb.sheetnames:
            return self.wb.create_sheet(title=lab_name)
        else:
            return self.wb.get_sheet_by_name(lab_name)

    def _get_section_ranges(self, sections):
        # skip lab name, total col
        column = 3
        ranges = {}
        for section in sections:
            start = column
            end = len(section["criteria"]) + start - 1
            ranges[section["name"]] = (start, end)
            column = end + 2
        return ranges

    def _setup_header(self, lab_config):
        self._active_sheet.cell(column=1,row=1).value = lab_config["lab name"]
        self._active_sheet.cell(column=2,row=1).value = "Total"

        sections_config = lab_config["sections"]
        for section in lab_config["sections"]:
            section_range = self._section_ranges[section["name"]]
            self._setup_section_header(section["criteria"], section_range)

    def _setup_section_header(self, section_config, section_range):
        for column, criteria in enumerate(section_config):
            cell = self._active_sheet.cell(column=column + section_range[0], row=1)
            cell.value = criteria["name"]
            cell.comment = openpyxl.comments.Comment("{}:{}".format(criteria["worth"], criteria["description"]), "Auto-Generator")

    def _create_group_rows(self, lab_config, groups, group_grades):
        for group_num in groups:
            self._create_group_row(lab_config, int(group_num) + 2, groups[int(group_num)], group_grades[group_num])

    def _create_group_row(self, lab_config, row_num, group, group_grade):
        self._active_sheet.cell(column=1, row=row_num).value = "Group {}".format(group.get_group_number())
        sum_val = self._build_sum_formula(lab_config["sections"], row_num)
        self._active_sheet.cell(column=2, row=row_num).value = sum_val

        self._add_auto_grades(lab_config["sections"], row_num, group_grade)

    def _build_sum_formula(self, sections, row_num):
        sums = []
        col_letter = openpyxl.utils.get_column_letter
        for section in sections:
            if section["include"]:
                start = col_letter(self._section_ranges[section["name"]][0])
                end = col_letter(self._section_ranges[section["name"]][1])

                sums.append("SUM({}{}:{}{})".format(start, row_num,
                                                   end, row_num))
        return "=MAX(0, " + " + ".join(sums) + ")"

    def _add_auto_grades(self, sections, row_num, group_grade):
        for section in sections:
            for i, criteria in enumerate(section["criteria"]):
                if "test" in criteria and criteria["test"] in group_grade:
                    test_name = criteria["test"]
                    column = self._section_ranges[section["name"]][0] + i
                    cell = self._active_sheet.cell(column=column, row=row_num)
                    grade = group_grade[test_name]
                    cell.value = grade[0] * criteria["possible"]
                    cell.comment = openpyxl.comments.Comment(str(grade[1]), "Auto-Grader")

    def get_grades(self, lab_config):
        lab_config = lab_config["assignment"]
        self._active_sheet = self._get_lab_sheet(lab_config["lab name"])
        self._section_ranges = self._get_section_ranges(lab_config["sections"])
        grades = self._load_group_grades(lab_config["sections"])
        return grades

    def _load_group_grades(self, sections):
        row = 2
        groups_grades = {}
        group_name = self._active_sheet.cell(row=row, column=1).value
        while group_name:
            grade = self._load_group_grade(sections, row)
            # TODO: Find better way to get group_num
            group_num = row - 2
            groups_grades[group_num] = grade

            row += 1
            group_name = self._active_sheet.cell(row=row, column=1).value
        return groups_grades

    def _load_group_grade(self, sections, group_row_num):
        group_num = group_row_num - 2
        group_row = self._active_sheet[group_row_num][2:]
        group_row = [cell for cell in group_row if cell.value]

        group_grade = {}
        comments = []
        grade = 0
        for section in filter(lambda sec: sec["include"], sections):
            section_range = self._section_ranges[section["name"]]
            for column, criteria in enumerate(section["criteria"]):
                column += section_range[0]
                cell = self._active_sheet.cell(column=column, row=group_row_num)
                cell_comment = self._get_comment(cell)
                points = int(cell.value)
                grade += points
                point_diff = int(points) - int(criteria["worth"])
                if point_diff != 0:
                    if not cell_comment:
                        raise ValueError("No note found on {}{}".format(cell.column, cell.row))
                    cell_comment = "{:+} : {}".format(point_diff, cell_comment)
                    comments.append(cell_comment)
        group_grade["Grade"] = grade
        group_grade["Note"] = "\n".join(filter(lambda x: x, comments))
        return group_grade

    def _get_comment(self, cell):
        com = cell.comment
        sub_comment = ""
        if com:
            sub_comment = com.text
        return sub_comment.strip()
