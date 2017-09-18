import openpyxl
import utils
from submissionChecks import GroupSubmissionGrader

class GradeSheet():
    def __init__(self, workbook_name):
        self.grading_column_amount = 0
        self.workbook_name = workbook_name
        self.wb = openpyxl.load_workbook(workbook_name)

    def create_grade_sheet(self, groups, rubric_name, lab_number):
        lab_name = "Lab " + str(lab_number)

        lab_sheet = None
        if lab_name not in self.wb.sheetnames:
            lab_sheet = self.wb.create_sheet(title=lab_name)
        else:
            lab_sheet = self.wb.get_sheet_by_name(lab_name)
        self._setup_header(lab_sheet, rubric_name, lab_name)
        self._create_group_rows(lab_sheet, groups)
        self.wb.save(self.workbook_name)

    def _setup_header(self, lab_sheet, rubric_name, lab_name):
        rubric = openpyxl.load_workbook(rubric_name)["Rubric"]
        penalities = self._get_penality_header(rubric)
        header = self._get_lab_header(rubric, lab_name)
        auto_header = self._get_auto_header(rubric)
        self._set_lab_header(lab_sheet, header, penalities, auto_header)

    def _get_lab_header(self, sheet, lab_name):
        for row_cell in sheet["A"]:
            if row_cell.value == lab_name:
                return [cell for cell in sheet[row_cell.row] if cell.value != None]

    def _get_penality_header(self, sheet):
        for row_cell in sheet["A"]:
            if row_cell.value == "Penalties":
                return [cell for cell in sheet[row_cell.row][1:] if cell.value != None]

    def _get_auto_header(self, sheet):
        for row_cell in sheet["A"]:
            if row_cell.value == "Auto-Info":
                return [cell for cell in sheet[row_cell.row][1:] if cell.value != None]

    def _set_lab_header(self, lab_sheet, lab_header, penality_header, auto_header):
        lab_sheet.cell(column=1,row=1).value = lab_header[0].value
        lab_sheet.cell(column=2,row=1).value = "Total"
        column = 3

        start = column
        self.grading_column_amount = len(lab_header[1:])
        for col_cell in lab_header[1:]:
            cell = lab_sheet.cell(column=column, row=1)
            cell.value = col_cell.value
            cell.comment = col_cell.comment
            column += 1
        self.grading_columns = (start, column - 1)

        column += 1
        start = column
        for col_cell in penality_header:
            cell = lab_sheet.cell(column=column, row=1)
            cell.value = col_cell.value
            cell.comment = col_cell.comment
            column += 1
        self.penality_columns = (start, column - 1)

        column += 1
        start = column
        for col_cell in auto_header:
            cell = lab_sheet.cell(column=column, row=1)
            cell.value = col_cell.value
            cell.comment = col_cell.comment
            column += 1
        self.auto_columns = (start, column - 1)
        column += 1

    def _create_group_rows(self, lab_sheet, groups):
        for group_num in groups:
            self._create_group_row(lab_sheet, int(group_num) + 2, groups[int(group_num)])

    def _create_group_row(self, lab_sheet, row_num, group):
        lab_sheet.cell(column=1, row=row_num).value = "Group {}".format(group.get_group_number())
        col_letter = openpyxl.utils.get_column_letter
        sum_val = "=SUM({}{}:{}{}) - SUM({}{}:{}{})".format(col_letter(self.grading_columns[0]), row_num,
                                           col_letter(self.grading_columns[1]), row_num,
                                           col_letter(self.penality_columns[0]), row_num,
                                           col_letter(self.penality_columns[1]), row_num)
        lab_sheet.cell(column=2, row=row_num).value = sum_val

        for column in range(self.penality_columns[0], self.penality_columns[1] + 1):
            cell = lab_sheet.cell(column=column, row=row_num)
            cell.value = 0

        column = self.auto_columns[0]
        group_grades = GroupSubmissionGrader().check_group(group)
        temp = self._extract_pep8(group_grades)
        temp = self._extract_syntax(group_grades)
        for extract_func in [self._extract_pep8, self._extract_syntax]:
            cell = lab_sheet.cell(column=column, row=row_num)
            grade = extract_func(group_grades)
            cell.value = grade[0]
            cell.comment = openpyxl.comments.Comment(str(grade[1]), "Auto-Grader")
            column += 1

    def _extract_pep8(self, group_grades):
        errors = group_grades["Pep8"]
        return (errors[0], "\n".join(errors[1]))

    def _extract_partner_grade(self, group_grades):
        # Needs to be per-student basis
        return ("-", None)

    def _extract_syntax(self, group_grades):
        compiles = group_grades["Syntax"]
        return (int(compiles), "Compilation " + ["Failed", "Completed"][compiles])

if __name__ == "__main__":
    submissions = utils.load_json("students.json")
    submissions = utils.convert_submission_dict_to_classes(submissions)
    groups = utils.load_json("groups.json")
    groups = utils.generate_group_submissions(submissions, groups)
    lab_number = 10
    rubric_name = "rubric.xlsx"
    lab_sheet_name = "grading-sheet.xlsx"
    GradeSheet(lab_sheet_name).create_grade_sheet( groups, rubric_name, lab_number)
